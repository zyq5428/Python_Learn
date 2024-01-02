#blankRowInserter.py
import sys, openpyxl
from openpyxl.utils import get_column_letter
#获取行号以及行数
lineRow, numRow = int(sys.argv[1]), int(sys.argv[2])
#打开源工作表，得到最大行、列
wbSour = openpyxl.load_workbook(sys.argv[3])
sheetSour = wbSour['Sheet']
maxRowSour,maxColumnSour = sheetSour.max_row, sheetSour.max_column

#打开目标工作表，使用get_column_letter得到要合并范围
wbAim = openpyxl.Workbook()
sheetAim = wbAim['Sheet']
numFirst = get_column_letter(1) + str(lineRow)
numLast = get_column_letter(maxColumnSour) + str(lineRow+numRow-1)
#行号前照抄，行号+行数来跳行
for i in range(1, maxRowSour + 1):
    for j in range(1, maxColumnSour + 1):
        if i < lineRow:
            sheetAim.cell(row=i, column=j).value = sheetSour.cell(row=i, column=j).value
        else:
            sheetAim.cell(row=i+numRow, column=j).value = sheetSour.cell(row=i, column=j).value
#将numFirst:+numLast的范围合并
sheetAim.merge_cells(''+numFirst+':'+numLast+'')

#保存
wbAim.save('text12311.xlsx') 
