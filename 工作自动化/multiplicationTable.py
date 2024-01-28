import openpyxl,sys
from openpyxl.styles import Font #输入Font代替输入openpyxl.styles.Font()

newExcel = openpyxl.Workbook() #新建工作簿对象
# wb = openpyxl.load_workbook("newExcel.xlsx")
# 表格输出
sheet = newExcel.get_sheet_by_name("Sheet")

if len(sys.argv) != 2:
    print("程序提示: python multiplicationTable.py n")
    sys.exit(1)
else:
    n = int(sys.argv[1]) #将sys.argv接收到的参数转换为int型

fontObj = Font(bold = True) #设置字体加粗

for i in range(1,n+1):
    sheet.cell(row=i+1,column = 1).value = i #对i+1行,1列进行赋值
    sheet.cell(column = i+1,row = 1).value = i #对1行，i+1列进行赋值
    sheet.cell(row=i + 1, column=1).font = fontObj #加粗
    sheet.cell(column=i + 1, row=1).font = fontObj

# 公式输出
for k in range(2,n+2):
    for p in range(2,n+2):
      sheet.cell(row =k,column =p).value = (k-1) * (p-1)

newExcel.save("newExcel.xlsx")
# wb.save("newExcel.xlsx")


# 程序需要做以下事情：

#能够从cmd命令行窗口接收参数 n
#运行python脚本，生成新的excel文件
#excel文件包含n*n的乘法矩阵

# 代码需要做一下事情：

#导入openpyxl,sys模块
#openpyxl.Workbook()创建新的工作薄对象
#get_sheet_by_name取得工作表对象
#Font()创建字体样式对象
#for循环进行乘法矩阵坐标系设置
#for循环进行乘法结果单元格填充
#save()方法保存
