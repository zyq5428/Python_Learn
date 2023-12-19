import openpyxl,sys
from openpyxl.styles import Font #输入Font代替输入openpyxl.styles.Font()
newExcel = openpyxl.Workbook() #新建工作簿对象

# 表格输出
sheet = newExcel.get_sheet_by_name("Sheet")
n = int(sys.argv[1]) #将sys.argv接收到的参数转换为int型
fontObj = Font(bold = True) #设置字体加粗
for i in range(1,n+1):
    sheet.cell(row=i+1,column = 1).value = i #对i+1行,1列进行赋值
    sheet.cell(column = i+1,row = 1).value = i #对1行，i+1列进行赋值
    sheet.cell(row=i + 1, column=1).font = fontObj #加粗
    sheet.cell(column=i + 1, row=1).font = fontObj

# 公式输出
for k in range(2,n+2):
    for p in range(2,n+2):
      sheet.cell(row =k,column =p).value = (k-1) * (p-1)

newExcel.save("newExcel.xlsx")
